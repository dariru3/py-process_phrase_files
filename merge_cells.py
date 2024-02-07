from openpyxl import load_workbook
from openpyxl.utils import range_boundaries

# Load an Excel workbook and select a worksheet
wb = load_workbook('path_to_your_excel_file.xlsx')
sheet = wb['Sheet1']  # Replace 'Sheet1' with the name of your target sheet

# Iterate through the merged cells in the worksheet
for merge_range in sheet.merged_cells.ranges:
    min_col, min_row, max_col, max_row = range_boundaries(str(merge_range))
    # Check if the merge is in the second column (B)
    if min_col == 2 and max_col == 2:
        # Apply the same merge to the first column (A)
        sheet.merge_cells(start_row=min_row, start_column=1, end_row=max_row, end_column=1)

# Save the workbook
wb.save('path_to_your_modified_excel_file.xlsx')